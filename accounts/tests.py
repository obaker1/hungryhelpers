from django.test import TestCase
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.urls import reverse
from django.contrib.auth.models import Permission
from findLocation.models import GoogleMapsResponse
from openpyxl import load_workbook

class SignUpTest(TestCase):
    def setUp(self):
        self.username = 'test'
        self.password = '>pve_hm*N*&x<qbP8u'
        self.first_name = 'Bob'
        self.last_name = 'Shanghai'
        self.email = 'test@example.com'

    def test_signup_page(self):
        # access signup page
        response = self.client.get('/accounts/signup/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify signup.html is being used
        self.assertTemplateUsed(response, template_name='registration/signup.html')

    def test_signup_form(self):
        # send signup data to page
        response = self.client.post('/accounts/signup/', data={
            'username': self.username,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'email': self.email,
            'password1': self.password,
            'password2': self.password
        }, follow=True)
        # check that page successfully loaded: status code = 200
        self.assertEqual(response.status_code, 200)
        # check if account was successfully made
        users = get_user_model().objects.all()
        self.assertEqual(users.count(), 1)


class LogInTest(TestCase):
    def setUp(self):
        # valid credentials
        self.credentials = {
            'username': 'test',
            'password': '>pve_hm*N*&x<qbP8u'}
        # invalid credentials
        self.credentials2 = {
            'username': 'test',
            'password': '>pve_hm*N&x<qbP8u'
        }
        User.objects.create_user(**self.credentials)

    def test_login_page(self):
        # access signup page
        response = self.client.get('/accounts/login/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify signup.html is being used
        self.assertTemplateUsed(response, template_name='registration/login.html')

    def test_login_success(self):
        # send login data
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        # check site redirection destination returns status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # check that the user successfully logged in
        self.assertTrue(response.context['user'].is_active)
        # check that the user has been redirected to home
        self.assertTemplateUsed(response, template_name='home.html')

    def test_login_fail(self):
        # send login data
        response = self.client.post('/accounts/login/', self.credentials2, follow=True)
        # check site redirection destination returns status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # check that the user was unsuccessful when logging in
        self.assertFalse(response.context['user'].is_active)
        # check that the user is shown appropriate message
        self.assertTrue('Please enter a correct username and password' in str(response.content))

class DashboardPageContentTest(TestCase):
    def setUp(self):
        # valid credentials
        self.credentials = {
            'username': 'test',
            'password': '>pve_hm*N*&x<qbP8u'
        }
        self.aboutUsMsg = "Since the COVID-19 pandemic, the distribution of subsidized meals has become an increasing"
        self.dashboardMsg = "Set up your pick up or delivery method for your meal plans"
        User.objects.create_user(**self.credentials)

    def test_dashboard_page_while_logged_off(self):
        # Go to homepage to load dashboard
        response = self.client.get('')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify home.html is being used
        self.assertTemplateUsed(response, template_name='home.html')
        # verify correct message is displayed to anonymous user
        self.assertContains(response, self.aboutUsMsg)

    def test_dashboard_page_while_logged_in(self):
        # send login data
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        # check site redirection destination returns status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # check that the user successfully logged in
        self.assertTrue(response.context['user'].is_active)
        # check that the user has been redirected to home
        self.assertTemplateUsed(response, template_name='home.html')
        # verify correct message is displayed to logged in user
        self.assertContains(response, self.dashboardMsg)

class LogOutTest(TestCase):
    def setUp(self):
        self.credentials = {
            'username': 'test',
            'password': '>pve_hm*N*&x<qbP8u'}
        User.objects.create_user(**self.credentials)

    def test_logout_page(self):
        # access logout page
        response = self.client.get('/accounts/logout/')
        # verify site redirection status code (HTTP 302)
        self.assertEqual(response.status_code, 302)

        # re-access logout page and follow redirection
        response = self.client.get('/accounts/logout/', follow=True)
        # verify site redirection destination returns status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify home.html is being used
        self.assertTemplateUsed(response, template_name='home.html')

    def test_logout(self):
        # send login data
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        # check that a page redirection occurred (HTTP 302)
        self.assertEqual(response.status_code, 200)
        # check that the user successfully logged in
        self.assertTrue(response.context['user'].is_active)
        # check that the user has been redirected to home
        self.assertTemplateUsed(response, template_name='home.html')
        # request logout
        response = self.client.get("/accounts/logout/", follow=True)
        # check site status code (HTTP 200 OK)
        self.assertEquals(response.status_code, 200)
        # check user is no longer active
        self.assertFalse(response.context['user'].is_active)
        # check user is shown "You are not logged in" message
        self.assertTrue('You are not logged in' in str(response.content))

class EditSettingsTest(TestCase):
    def setUp(self):
        # valid credentials
        self.credentials = {
            'username': 'test',
            'password': '>pve_hm*N*&x<qbP8u'}
        User.objects.create_user(**self.credentials)

        # information for settings page
        self.username = 'test'
        self.email = 'tester@tset.com'
        self.first_name = 'testing'
        self.last_name = 'mctest'

    def test_settings_page_while_logged_out(self):
        # access settings page
        response = self.client.get('/accounts/settings/', follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify a redirect was issued to take user to the login page
        self.assertTemplateUsed(response, template_name='registration/login.html')

    def test_settings_page_while_logged_in(self):
        # send login data
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        # check site redirection destination returns status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # check that the user successfully logged in
        self.assertTrue(response.context['user'].is_active)
        # check that the user has been redirected to home
        self.assertTemplateUsed(response, template_name='home.html')
        # access settings page
        response = self.client.get('/accounts/settings/', follow=True)
        self.assertEqual(response.status_code, 200)
        # verify edit_settings.html is being used
        self.assertTemplateUsed(response, template_name='registration/edit_settings.html')
        # send and save information to settings page
        response = self.client.post('/accounts/settings/', data={
            'username': self.username,
            'email': self.email,
        }, follow=True)
        # check that the user has been redirected to home after updating settings
        self.assertTemplateUsed(response, template_name='home.html')
        # reaccess settings page
        response = self.client.get('/accounts/settings/')
        # verifies that the information has been successfully updated (username, email, fn, and ln)
        self.assertContains(response, self.username)
        self.assertContains(response, self.email)

class ProfileTest(TestCase):
    def setUp(self):
        # valid credentials
        self.username = 'test'
        self.password = '>pve_hm*N*&x<qbP8u'
        self.email = 'test@example.com'
        self.parent_first_name = "Mr"
        self.parent_last_name = "Parent"

        # information for edit parent page
        self.address = '1000 Hilltop Cir'
        self.city = 'Baltimore'
        self.state = 'MD'
        self.zip = '21250'
        self.district = 'Baltimore County'

        # information for add student page
        self.first_name = 'Billy'
        self.last_name = 'Manson'
        self.new_last_name = 'YummyPaint'
        self.age = 11
        self.school = 'Catonsville High'
        self.grade = 8
        self.student_id = 'AB04576'
        self.allergic_celiac = 'Yes'
        self.allergic_shellfish = 'No'
        self.allergic_lactose = 'Yes'
        self.preference_halal = 'Yes'
        self.preference_kosher = 'No'
        self.preference_vegetarian = 'No'

        # information for meal plan
        self.pickup_type = 'School'
        self.day = 'M/W'
        self.time = '11:00am-11:15am'
        self.meal_breakfast = "Yes"
        self.meal_lunch = "No"
        self.meal_dinner = "No"
        self.pickup_location = "Westland Gardens Apartments: 2.0 miles in 5 mins (11:00am-12:30pm)"
        self.new_pickup_location = "Catonsville High: 1.5 miles in 5 mins (M/W 11am-1pm)"
        self.complete = "Yes"

        self.err_msg = "You are either not logged in or do not have access to this profile."

        LOCATION = 1; DISTANCE = 2; TIME = 3; BUS = 4; SCHOOL = 5; ADDRESS = 6; TIMEFRAME = 7; LATITUDE = 8; LONGITUDE = 9;
        file = "Locations.xlsx"
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        for data in sheet.iter_rows(values_only=True):
            newDest = GoogleMapsResponse(location=data[LOCATION], distance=data[DISTANCE],
                                         time=data[TIME], bus=data[BUS], school=data[SCHOOL],
                                         address=data[ADDRESS], timeframe=data[TIMEFRAME],
                                         latitude=data[LATITUDE], longitude=data[LONGITUDE])
            newDest.save()

        #loc = GoogleMapsResponse(location="Arbutus Middle", distance=2.9, time="7 mins", school='F', bus='T', timeframe='M/W 11am-1pm', latitude=39.248289, longitude=-76.7057813)

    def test_profile_page(self):
        """ sign up and log into account """
        # access and create account on signup page
        response = self.client.post('/accounts/signup/', data={
            'username': self.username,
            'first_name': self.parent_first_name,
            'last_name': self.parent_last_name,
            'email': self.email,
            'password1': self.password,
            'password2': self.password
        }, follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # check if account was successfully made
        users = get_user_model().objects.all()
        self.assertEqual(users.count(), 1)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify redirection was successfully made to home
        self.assertTemplateUsed(response, template_name='home.html')

        """ access profile page """
        # access profile page
        response = self.client.get('/accounts/profile/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/profile.html')

        """ access the meal plan page before updating location in profile """
        # access meal plan page
        response = self.client.get('/accounts/meal_plans/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/meal_plans.html')
        # verify that user is shown "Create Meal Plan" button
        self.assertContains(response, "Please update your profile with your location before using this feature!")

        """ access the edit profile page """
        # access profile page
        response = self.client.get('/accounts/edit_profile/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/edit_profile.html')

        """ edit profile content """
        # access profile page
        response = self.client.post('/accounts/edit_profile/', data={
            'address': self.address,
            'city': self.city,
            'state': self.state,
            'zip': self.zip,
            'district': self.district,
        }, follow=True)
        # verify success of form submission
        self.assertEqual(response.status_code, 200)
        # verify correct template was used
        # After successful submission, user should be taken back to profile page
        self.assertTemplateUsed(response, template_name='registration/profile.html')
        self.assertContains(response, self.parent_first_name)
        self.assertContains(response, self.parent_last_name)
        self.assertContains(response, self.city)
        self.assertContains(response, self.state)
        self.assertContains(response, self.zip)
        self.assertContains(response, self.district)

        """ add a new student """
        # access page that allows caretaker to add student
        response = self.client.get('/accounts/add_student/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/add_student.html')
        # enter data into form

        response = self.client.post('/accounts/add_student/', data={
            'first_name': self.first_name,
            'last_name': self.last_name,
            'age': self.age,
            'grade': self.grade,
            'school' : self.school,
            'student_id' : self.student_id,
            'allergic_celiac' : self.allergic_celiac,
            'allergic_shellfish' : self.allergic_shellfish,
            'allergic_lactose' : self.allergic_lactose,
            'preference_halal' : self.preference_halal,
            'preference_kosher' : self.preference_kosher,
            'preference_vegetarian' : self.preference_vegetarian,
        }, follow=True)

        # verify success of form submission
        self.assertEqual(response.status_code, 200)
        # verify correct template was used
        # After successful submission, user should be taken back to profile page
        self.assertTemplateUsed(response, template_name='registration/profile.html')
        # verify student information appears on page
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.last_name)
        self.assertContains(response, self.school)
        self.assertContains(response, self.student_id)

        """ access the edit_student page for the newly created student profile"""
        response = self.client.get("/accounts/1/edit_student", follow=True)
        # check site status code (HTTP 200 OK)
        self.assertEquals(response.status_code, 200)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/edit_student.html')

        """ access the edit_student page for the newly created student profile and make edits"""
        response = self.client.post('/accounts/1/edit_student/', data={
            'first_name': self.first_name,
            'last_name': self.new_last_name,
            'age': self.age,
            'school' : self.school,
            'grade' : self.grade,
            'student_id' : self.student_id,
            'allergic_celiac' : self.allergic_celiac,
            'allergic_shellfish' : self.allergic_shellfish,
            'allergic_lactose' : self.allergic_lactose,
            'preference_halal' : self.preference_halal,
            'preference_kosher' : self.preference_kosher,
            'preference_vegetarian' : self.preference_vegetarian,
        }, follow=True)
        self.assertEqual(response.status_code, 200)
        # verify correct template was used
        # After successful submission, user should be taken back to profile page
        self.assertTemplateUsed(response, template_name='registration/profile.html')

        # returns to profile page to check changes
        response = self.client.get('/accounts/profile/', follow=True)
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        self.assertNotContains(response, self.last_name)
        self.assertContains(response, self.school)
        self.assertContains(response, self.student_id)

        """ access the meal plan page """
        # access meal plan page
        response = self.client.get('/accounts/meal_plans/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/meal_plans.html')
        # verify an element is shown with the child's first and last name
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        # verify that user is shown "Create Meal Plan" button
        self.assertContains(response, "Create Meal Plan")
        self.assertNotContains(response, "Edit Meal Plan")

        """ 'create' a meal plan for student """
        # access meal plan page
        response = self.client.get('/accounts/1/edit_meal_plan/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/edit_meal_plan.html')
        # verify meal plan is specific for the child
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        # verify certain information and fields are hidden until initial form is submitted
        self.assertNotContains(response, "Current Pickup Location:")
        self.assertNotContains(response, "Select Pickup Location")
        self.assertNotContains(response, "Meal Plan is ready to be seen by staff")
        # fill out meal plan form
        response = self.client.post('/accounts/1/edit_meal_plan/', data={
            'pickup_type': self.pickup_type,
            'day': self.day,
            'time': self.time,
            'meal_breakfast' : self.meal_breakfast,
            'meal_lunch': self.meal_lunch,
            'meal_dinner': self.meal_dinner,
        }, follow=True)
        self.assertTemplateUsed(response, template_name='registration/edit_meal_plan.html')
        # verify that previously hidden information and fields are now visible
        self.assertContains(response, "Current Pickup Location:")
        self.assertContains(response, "Select Pickup Location")
        self.assertContains(response, "Meal Plan is ready to be seen by staff")
        # verify that user is shown correct pickup locations relative to the location set in profile
        self.assertContains(response, "Westland Gardens Apartments: 2.0 miles in 5 mins (11:00am-12:30pm)")
        self.assertContains(response, "Catonsville High: 1.5 miles in 5 mins (M/W 11am-1pm)")
        self.assertContains(response, "1037 Maiden Choice Lane (Kendale Apartments): 2.0 miles in 5 mins (11:50am-12:30pm)")
        # fill out the rest of the meal plan
        response = self.client.post('/accounts/1/edit_meal_plan/', data={
            'pickup_type': self.pickup_type,
            'day': self.day,
            'time': self.time,
            'meal_breakfast' : self.meal_breakfast,
            'meal_lunch': self.meal_lunch,
            'meal_dinner': self.meal_dinner,
            'pickup_location': self.pickup_location,
            'complete': self.complete,
        }, follow=True)
        # verify that user was taken back to the same page
        self.assertTemplateUsed(response, template_name='registration/edit_meal_plan.html')

        """ access the meal plan page after creating a meal plan """
        # access meal plan page
        response = self.client.get('/accounts/meal_plans/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/meal_plans.html')
        # verify an element is shown with the child's first and last name
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        # verify that user is shown "Edit Meal Plan" button instead of "Create Meal Plan"
        self.assertNotContains(response, "Create Meal Plan")
        self.assertContains(response, "Edit Meal Plan")
        # verify that user is shown information concerning the meal plan
        self.assertContains(response, self.pickup_type)
        self.assertContains(response, self.time)
        self.assertContains(response, self.pickup_location)
        self.assertContains(response, "Complete and viewable by staff")

        """ edit existing meal plan for same student """
        # access meal plan page
        response = self.client.get('/accounts/1/edit_meal_plan/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/edit_meal_plan.html')
        # verify meal plan is specific for the child and all fields are visible
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        self.assertContains(response, "Current Pickup Location:")
        self.assertContains(response, "Select Pickup Location")
        self.assertContains(response, "Meal Plan is ready to be seen by staff")
        # verify that user is shown correct pickup locations relative to the location set in profile
        self.assertContains(response, "Westland Gardens Apartments: 2.0 miles in 5 mins (11:00am-12:30pm)")
        self.assertContains(response, "Catonsville High: 1.5 miles in 5 mins (M/W 11am-1pm)")
        self.assertContains(response, "1037 Maiden Choice Lane (Kendale Apartments): 2.0 miles in 5 mins (11:50am-12:30pm)")
        # change pickup location
        response = self.client.post('/accounts/1/edit_meal_plan/', data={
            'pickup_type': self.pickup_type,
            'day': self.day,
            'time': self.time,
            'meal_breakfast' : self.meal_breakfast,
            'meal_lunch': self.meal_lunch,
            'meal_dinner': self.meal_dinner,
            'pickup_location': self.new_pickup_location,
            'complete': "No",
        }, follow=True)
        # verify that user was taken back to the same page
        self.assertTemplateUsed(response, template_name='registration/edit_meal_plan.html')

        """ access the meal plan page after editing meal plan """
        # access meal plan page
        response = self.client.get('/accounts/meal_plans/', follow=True)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/meal_plans.html')
        # verify an element is shown with the child's first and last name
        self.assertContains(response, self.first_name)
        self.assertContains(response, self.new_last_name)
        # verify that user is shown "Edit Meal Plan" button instead of "Create Meal Plan"
        self.assertNotContains(response, "Create Meal Plan")
        self.assertContains(response, "Edit Meal Plan")
        # verify that user is shown information concerning the meal plan
        self.assertContains(response, self.pickup_type)
        self.assertContains(response, self.time)
        self.assertContains(response, self.new_pickup_location)
        self.assertContains(response, "Incomplete")

        """ delete the student and check that student information is no longer existent on profile page """
        response = self.client.get("/accounts/1/delete_student", follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify correct template was used
        self.assertTemplateUsed(response, template_name='registration/delete_student.html')

        # delete student from database
        response = self.client.post('/accounts/1/delete_student/', follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # After successful deletion, user should be taken back to profile page
        self.assertTemplateUsed(response, template_name='registration/profile.html')

        # returns to profile page to check changes
        response = self.client.get('/accounts/profile/', follow=True)
        self.assertNotContains(response, self.first_name)
        self.assertNotContains(response, self.new_last_name)
        self.assertNotContains(response, self.school)
        self.assertNotContains(response, self.student_id)

        """ log out of user's current session and attempt to access website pages """
        response = self.client.get("/accounts/logout/", follow=True)
        # check site status code (HTTP 200 OK)
        self.assertEquals(response.status_code, 200)
        # check user is no longer active
        self.assertFalse(response.context['user'].is_active)
        # check user is shown "You are not logged in" message
        self.assertContains(response, 'You are not logged in')
        # attempt to access profile page
        response = self.client.get('/accounts/profile/', follow=True)
        # verify that the user was redirected to login page
        self.assertTemplateUsed(response, template_name='registration/login.html')

        """ attempt to access edit page for student profile while logged out """
        response = self.client.get('/accounts/1/edit_profile/', follow=True)
        # verify that user is met with a 404 site status code
        self.assertEqual(response.status_code, 404)


class PasswordResetTest(TestCase):

    def setUp(self):
        # valid credentials
        self.credentials = {
            'username': 'test',
            'password': '>pve_hm*N*&x<qbP8u'}

        # valid credentials
        self.username = 'test'
        self.email = 'myemail@test.com'
        self.password = '>pve_hm*N*&x<qbP8u'
        self.new_password = '>pve_hm*N*&x<qbP8sss'
        # information for forgot password
        self.email = 'tester@tset.com'

        # create user
        self.the_user = User.objects.create_user(self.username, self.email, self.password)




    def test_password_change_form(self):
        # Login user
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        response = self.client.get('/accounts/password_change/', follow=True)
        self.assertTemplateUsed(response, template_name='registration/password_change_form.html')


        # access password change page
        response = self.client.post('/accounts/password_change/', data={
            'old_password': self.password,
            'new_password1' : self.new_password,
            'new_password2': self.new_password,
        }, follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_change.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_change_done.html')

    def test_password_change_done(self):
        response = self.client.post('/accounts/login/', self.credentials, follow=True)
        # access password change done page
        response = self.client.get('/accounts/password_change/done/', follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_change_done.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_change_done.html')


    def test_password_reset_form(self):
        # access password reset page and enter email
        response = self.client.post('/accounts/password_reset/', data={
            'email': self.email
        }, follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_reset_done.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_reset_done.html')

    def test_password_reset_done(self):
        # access password reset done page
        response = self.client.get('/accounts/password_reset/done/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_reset_done.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_reset_done.html')

    def test_password_reset_confirm(self):
        # access reset password page and allow user to enter new passwords
        response = self.client.post('/accounts/reset/<uidb64>/<token>/', data={
            'password1': self.password,
            'password2': self.password
        }, follow=True)
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_reset_confirm.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_reset_confirm.html')

    def test_password_reset_complete(self):
        # access password reset complete page
        response = self.client.get('/accounts/reset/done/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify password_reset_complete.html is being used
        self.assertTemplateUsed(response, template_name='registration/password_reset_complete.html')

class PermissionsTest(TestCase):
    fixtures = ['dbcontent.json', ]
    def test_admin_permissions(self):
        # login to admin user
        username = 'admin'
        password = 'admin'

        # Login to correct user
        self.client.login(username=username, password=password)

        # Open page and ensure that admin has access to add a location
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Map Configuration Panel")

    def test_staff_permissions(self):
        permission = Permission.objects.get(name='Can add google maps response')

        # create staff user
        password = 'staff'
        my_staff = User.objects.create_user('staff', 'myemail@test.com', password)

        # You'll need to log him in before you can send requests through the client
        self.client.login(username=my_staff.username, password=password)
        my_staff.user_permissions.add(permission)

        # Open page and ensure that staff also has access to add a location
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Map Configuration Panel")

    def test_parent_permissions(self):
        # create parent user
        password = 'parent'
        my_parent = User.objects.create_user('parent', 'myemail@test.com', password)

        # You'll need to log him in before you can send requests through the client
        self.client.login(username=my_parent.username, password=password)

        # Open page and ensure that staff also has access to add a location
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Map Configuration Panel")
