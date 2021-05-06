from django.test import TestCase, Client
from django.urls import reverse
import requests
from django.conf import settings
from django.contrib.auth.models import User
from .models import Origin, GoogleMapsResponse
from accounts.models import Profile
from openpyxl import load_workbook
from django.contrib.auth.models import Permission

# Create your tests here.
class NoDatabase(TestCase):
    # test findLocation page works even when no database is loaded
    def test_page_load(self):
        # access findLocation page
        response = self.client.get('/findLocation/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify index.html is being used
        self.assertTemplateUsed(response, template_name='findLocation/index.html')

class OriginIndexViewTest(TestCase):
    #fixtures = ['dbcontent.json', ]
    # test adding origin
    def test_adding_origin(self):
        c = Client()
        # put default origins and destinations
        c.post('/findLocation/addOrigin/', {'origin': 'baltimore, MD'})
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)

class DestinationIndexViewTest(TestCase):
    #fixtures = ['dbcontent.json', ]
    def setUp(self):
        self.username = 'admin'
        self.first_name = 'admin'
        self.last_name = 'user'
        self.email = 'tester@test.com'
        self.password = '>pve_hm*N*&x<qbP8u'

        User.objects.create_superuser(self.username, self.email, self.password)
        admin = User.objects.get(username=self.username)
        admin.first_name, admin.last_name = 'admin', 'user'
        admin.save()
        profile = Profile(user=admin, address='1000 Hilltop Cir', city='Baltimore', state='MD', zip='21250',
                          district='Baltimore County')
        profile.save()
        origin = Origin(user=admin, origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213,
                        longitude=-76.7143524)
        origin.save()

    def test_adding_destination(self):
        """
        Make sure that a destination inputted is in the database.
        """
        response = self.client.post('/accounts/login/', data={
            'username': self.username,
            'password': self.password,
        }, follow=True)


        c = Client()
        c.post('/findLocation/addOrigin/', {'origin': 'baltimore, MD'})
        c.post('/findLocation/addLocation/', {'destination': 'Baltimore Avenue and 5th Avenue', 'school': 'y', 'bus': 'n', 'timeframe': 'M\W', 'remove': 'a'})
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Baltimore Avenue and 5th Avenue")

    def test_removing_destination(self):
        """
        Make sure that a destination is removed from the database.
        """
        c = Client()
        # adding a location
        c.post('/findLocation/addLocation/', {'destination': 'Towson, Maryland', 'school': 'y', 'bus': 'n', 'timeframe': 'M\W', "remove": 'a', 'origin': 'Towson MD'})
        # removing a location
        c.post('/findLocation/addLocation/', {'destination': 'Towson, Maryland', 'school': 'y', 'bus': 'n', 'timeframe': 'M\W', "remove": 'r', 'origin': 'Towson MD'})
        response = self.client.get(reverse('findlocation'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Towson,")

    def test_google_matrix_api(self):
        """
        makes sure that distance matrix API is working
        """
        params = {
            'key': settings.GOOGLE_MAPS_API_KEY,
            'origins': ['Columbia, MD'],
            'destinations': 'Towson, Maryland'
        }
        # calls distance matrix API
        url = 'https://maps.googleapis.com/maps/api/distancematrix/json?callback=initMap&libraries=&v=weekly&units=imperial'
        response = requests.get(url, params)
        origin_address = response.json()['origin_addresses'][0]
        self.assertIn('Columbia, MD', origin_address)
        self.assertEqual(response.status_code, 200)

    def test_google_geocode_api(self):
        """
        makes sure that geocode API is working
        """
        url = 'https://maps.googleapis.com/maps/api/js?'
        response = requests.get(url)
        self.assertEqual(response.status_code, 200)

class TestOrdering(TestCase):
    def test_order_of_locations(self):
        """
        makes sure that all locations in the database are ordered
        correctly by distance (not by time)
        """
        c = Client()
        newOrigin = Origin(origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213, longitude=-76.7143524)
        newOrigin.save()
        destinationList = ['Elkridge, MD', 'Towson, MD', 'Columbia, MD' ]
        destinationListCorrect = ['Elkridge, MD', 'Columbia, MD', 'Towson, MD']
        for i in destinationList:
            c.post('/findLocation/addLocation/', {'destination': i, 'school': 'y', 'bus': 'n', 'timeframe': 'M\W', "remove": 'a', 'origin': 'Towson MD'})
            response = self.client.get(reverse('findlocation'))
        counter = 0
        googlemaps = GoogleMapsResponse.objects.all().order_by('distance', 'location')
        for destinations in googlemaps:
            self.assertIn(destinationListCorrect[counter], destinations.location)
            counter += 1

        self.assertEqual(response.status_code, 200)

class TestMoreLocations(TestCase):
    #fixtures = ['dbcontent.json', ]
    def setUp(self):
        file = "Locations.xlsx"
        workbook = load_workbook(filename=file)
        sheet = workbook.active

        for data in sheet.iter_rows(values_only=True):
            newDest = GoogleMapsResponse(location=data[1], distance=data[2],
                                         time=data[3], bus=data[4], school=data[5],
                                         address=data[6], timeframe=data[7],
                                         latitude=data[8], longitude=data[9])
            newDest.save()

    def test_show_more_locations(self):
        """
        Make sure page loads for showing 10 more locations
        """
        response = self.client.get('/findLocation/addMore/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify index.html is being used
        self.assertTemplateUsed(response, template_name='findLocation/index.html')

class staffPermissions(TestCase):
    def setUp(self):
        permission = Permission.objects.get(name='Can add google maps response')
        self.username = 'staff'
        self.first_name = 'staff'
        self.last_name = 'staff'
        self.email = 'tester@test.com'
        self.password = 'staff'

        # create staff user
        User.objects.create_user(self.username, self.email, self.password)
        my_staff = User.objects.get(username=self.username)
        my_staff.first_name, my_staff.last_name = self.first_name, self.last_name
        my_staff.save()
        profile = Profile(user=my_staff, address='1000 Hilltop Cir', city='Baltimore', state='MD', zip='21250',
                          district='Baltimore County')
        profile.save()
        origin = Origin(user=my_staff, origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213,
                        longitude=-76.7143524)
        origin.save()

        # You'll need to log him in before you can send requests through the client
        self.client.login(username=my_staff.username, password=self.password)
        my_staff.user_permissions.add(permission)

    def test_show_more_locations(self):
        """
        Make sure page loads for staff showing more locations
        """
        response = self.client.get('/findLocation/addMore/')
        # verify site status code (HTTP 200 OK)
        self.assertEqual(response.status_code, 200)
        # verify index.html is being used
        self.assertTemplateUsed(response, template_name='findLocation/index.html')