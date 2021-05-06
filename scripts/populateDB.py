def run():
    from findLocation.models import GoogleMapsResponse, Origin
    from accounts.models import Profile, Student, MealPlan
    from django.contrib.auth.models import User
    import os
    from openpyxl import load_workbook
    LOCATION = 1; DISTANCE = 2; TIME = 3; BUS = 4; SCHOOL = 5; ADDRESS = 6; TIMEFRAME = 7; LATITUDE = 8; LONGITUDE = 9;
    username = "admin"; email = "myemail@test.com"; password = "admin";
    reg_username = "user"; reg_password = "user";

    print("Flushing all tables in db.sqlite3")
    os.system("python manage.py flush --noinput")
    print("Running makemigrations...")
    os.system("python manage.py makemigrations")
    print("Running migrate...")
    os.system("python manage.py migrate")

    User.objects.create_superuser(username, email, password)
    admin = User.objects.get(username=username)
    admin.first_name, admin.last_name = 'admin', 'user'
    admin.save()
    profile = Profile(user=admin, address='1000 Hilltop Cir', city='Baltimore', state='MD', zip='21250', district='Baltimore County')
    profile.save()
    print("Created superuser account (username: " + username + ", password: " + password + ")")
    origin = Origin(user=admin, origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213, longitude=-76.7143524)
    origin.save()
    #print("Set origin for findLocation")

    User.objects.create_user(reg_username, email, reg_password)
    reg_user = User.objects.get(username=reg_username)
    reg_user.first_name, reg_user.last_name = 'End', 'User'
    reg_user.save()
    profile = Profile(user=reg_user, address='1000 Hilltop Cir', city='Baltimore', state='MD', zip='21250',
                      district='Baltimore County')
    profile.save()
    print("Created regular user account (username: " + reg_username + ", password: " + reg_password + ")")
    origin = Origin(user=reg_user, origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213,
                    longitude=-76.7143524)
    origin.save()
    student = Student(user_account=profile, first_name="Billy", last_name="Bob", age=7, grade=8, school="Catonsville High", student_id="AB04576",
                      allergic_celiac="No", allergic_shellfish="Yes", allergic_lactose="No",
                      preference_halal="Yes", preference_kosher="Yes", preference_vegetarian="No")
    student.save()
    mealplan = MealPlan(student_profile=student, pickup_type='', time='', day='', meal_breakfast='', meal_lunch='',
                        meal_dinner='')
    mealplan.save()
    #print("Set origin for findLocation")
    #newOrigin = Origin(origin='1000 Hilltop Cir, Baltimore, MD 21250, USA', latitude=39.2537213, longitude=-76.7143524)
    #newOrigin.save()


    file = "Locations.xlsx"
    workbook = load_workbook(filename=file)
    sheet = workbook.active

    for data in sheet.iter_rows(values_only=True):
        newDest = GoogleMapsResponse(location=data[LOCATION], distance=data[DISTANCE],
                                        time=data[TIME], bus=data[BUS], school=data[SCHOOL],
                                        address=data[ADDRESS], timeframe=data[TIMEFRAME],
                                        latitude=data[LATITUDE], longitude=data[LONGITUDE])
        newDest.save()

    print("Added all locations for findLocation")
    print("Launching webserver...")
    os.system("python manage.py runserver")